# -*- coding:utf-8 -*-

import datetime
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font


def CalSupposedPeople(firstPeopleNum, secondPeopleNum, supposedPeopleNum):
    temp = (firstPeopleNum + secondPeopleNum) / 2
    result = temp / supposedPeopleNum
    if result[0] >= 1.00:
        result = float(.99999)
    return result


def SetExcelStyle(path):
    wb = openpyxl.load_workbook(filePath)
    sheetName = wb.sheetnames
    # 自适应调整列宽做不出来，只能强行定义每一列的列宽
    columnName = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    widthProcessed_list = [13.11, 8, 13.11, 10.22, 10.22, 17.89, 17.89, 17.89, 17.89, 10.22, 10.22, 10.22, 17.89, 25, 13.11]
    widthSchool_list = [8.2, 17.89, 17.89, 17.89, 17.89, 17.89, 17.89, 17.89, 10.22, 9.67, 7.9, 8.9, 7.9, 7.9]
    widthGrouped_list = [8.1, 11, 11, 11, 11, 11]
    for name in sheetName:
        ws_temp = wb[name]
        if name == '数据汇总':
            for r in range(1, ws_temp.max_row + 1):
                if r == 1:
                    ws_temp.row_dimensions[r].height = 42.6  # 设置行高
                    ws_temp.merge_cells('A1:F1')
                    font = Font(u'楷体', size=20, bold=True, color='000000')
                elif r == 2:
                    ws_temp.row_dimensions[r].height = 30.6  # 设置行高
                    font = Font(u'楷体', size=14, bold=True, color='000000')
                    border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
                else:
                    ws_temp.row_dimensions[r].height = 30.6  # 设置行高
                    font = Font(u'楷体', size=12, color='FF000000')
                    border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
                alignment = Alignment(horizontal='center', vertical='center')
                for c in range(1, ws_temp.max_column + 1):
                    if r >= 3:
                        ws_temp.cell(row=r, column=c).number_format = '0.00%'
                    ws_temp.cell(row=r, column=c).font = font
                    ws_temp.cell(row=r, column=c).border = border
                    ws_temp.cell(row=r, column=c).alignment = alignment
                for index in range(len(widthGrouped_list)):
                    col_name = columnName[index]
                    ws_temp.column_dimensions[col_name].width = widthGrouped_list[index]
        elif name == '处理后数据':
            for r in range(1, ws_temp.max_row + 1):
                ws_temp.row_dimensions[r].height = 28.2  # 设置行高
                if r == 1:
                    border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
                else:
                    border = Border() # 选用默认情况就足够了
                font = Font(u'楷体', size=12, color='000000')
                alignment = Alignment(horizontal='center', vertical='center')
                for c in range(1, ws_temp.max_column+1):
                    ws_temp.cell(row=r, column=c).font = font
                    ws_temp.cell(row=r, column=c).border = border
                    ws_temp.cell(row=r, column=c).alignment = alignment
                for index in range(len(widthProcessed_list)):
                    col_name = columnName[index]
                    ws_temp.column_dimensions[col_name].width = widthProcessed_list[index]
        elif name == '原始数据':
            pass
        else:   # 学院的数据
            for r in range(1, ws_temp.max_row + 1):
                if r == 1 or r == 4:
                    ws_temp.row_dimensions[r].height = 31.2  # 设置行高
                    border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
                else:
                    ws_temp.row_dimensions[r].height = 15.6  # 设置行高
                    border = Border()  # 选用默认值
                font = Font(u'楷体', size=12, color='FF000000')
                alignment = Alignment(horizontal='center', vertical='center')
                for c in range(1, ws_temp.max_column+1):
                    if r == 2 and c >= 10:
                        ws_temp.cell(row=r, column=c).number_format = '0.00%'
                    if r >= 3 and c >= 11:
                        ws_temp.cell(row=r, column=c).border = Border()
                    else:
                        ws_temp.cell(row=r, column=c).font = font
                        ws_temp.cell(row=r, column=c).border = border
                        ws_temp.cell(row=r, column=c).alignment = alignment
                for index in range(len(widthSchool_list)):
                    col_name = columnName[index]
                    ws_temp.column_dimensions[col_name].width = widthSchool_list[index]
    wb.save(filePath)


def ProcessOriginExcel(path):
    file = pd.read_excel(path, sheet_name='处理后数据')
    originDataFrame = pd.DataFrame(file)
    schoolName = set(originDataFrame.iloc[:, 1])  # 读取第二列数据，为所有学院的名字
    schoolDataFrame_list = []  # 创建一个列表，用于储存所有学院单独的数据帧
    schoolProcessedDataFrame_list = []  # 创建一个列表，用于储存所有学院经过计算后的数据帧
    # 用于存储汇总后的每个学院数据表头
    groupedDataHead = ['学院', '应到人数', '第一次出勤人数', '第二次出勤人数',  # 前两列直接将现成的数据累加
                       '平均人数', '7:30迟到人数', '8:00早退人数', '违纪人数', '请假人数',
                       '迟到率', '早退率', '出勤率', '违纪率', '请假率']  # 这一列的数据需要程序进行计算
    # 数据按学院分类汇总
    for name in schoolName:
        tempSchoolData = originDataFrame.loc[(originDataFrame['学院'] == name)]  # 提取出每一个学院
        tempSchoolData.pop('工作日')
        groupedDataFrame = pd.DataFrame(np.arange(14).reshape((1, 14)),
                                        columns=groupedDataHead)  # 14对应的是groupedDataHead的列数
        groupedDataFrame.iloc[0, 0] = name
        for index in range(1, 9):  # 这里取的是groupedHead中直接将现成的数据累加的,比如应到人数、迟到人数等等
            tempHeadName = groupedDataHead[index]
            groupedDataFrame[tempHeadName] = tempSchoolData[tempHeadName].sum(axis=0)
        groupedDataFrame['迟到率'] = groupedDataFrame['7:30迟到人数'] / groupedDataFrame['平均人数']
        groupedDataFrame['早退率'] = groupedDataFrame['8:00早退人数'] / groupedDataFrame['平均人数']
        groupedDataFrame['出勤率'] = CalSupposedPeople(groupedDataFrame['第一次出勤人数'],
                                      groupedDataFrame['第二次出勤人数'], groupedDataFrame['应到人数'])
        groupedDataFrame['违纪率'] = groupedDataFrame['违纪人数'] / groupedDataFrame['平均人数']
        groupedDataFrame['请假率'] = groupedDataFrame['请假人数'] / groupedDataFrame['平均人数']
        # 数据存入一个列表，方便后续用openpyxl写入excel
        schoolDataFrame_list.append(tempSchoolData)
        schoolProcessedDataFrame_list.append(groupedDataFrame)

    wb = openpyxl.load_workbook(filePath)
    # 各个学院的数据写入excel中的每一个学院对应的sheet
    for index in range(len(schoolDataFrame_list)):
        name_temp = schoolDataFrame_list[index]['学院'].iat[0]  # 获取series数组的一个值，也就是学院的名称
        ws_temp = wb.create_sheet(title=name_temp)

        # 各个学院“计算后”的数据导入excel
        rows = dataframe_to_rows(schoolProcessedDataFrame_list[index])
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                if c_idx == 1:  # 不知道为什么第一列没有数据，要从第二列开始
                    pass
                else:
                    ws_temp.cell(row=r_idx, column=c_idx - 1, value=value)
        ws_temp.delete_rows(2, 1)  # 删除一个空行，我也不知道为什么会出现一个空行

        # 各个学院的数据导入excel
        rows = dataframe_to_rows(schoolDataFrame_list[index])
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                if c_idx != 1:
                    ws_temp.cell(row=r_idx + 3, column=c_idx - 1, value=value)
                if c_idx == 11 or r_idx == 2:  # 后面的备注就不导入进去了,并且跳过一个空行
                    break
        ws_temp.delete_rows(5, 1)  # 删除一个空行，我也不知道为什么会出现一个空行

    # 将每个学院汇总的信息整理出来放到一个新的sheet
    ws_temp = wb.create_sheet(index=2, title='数据汇总')  # 在excel中的第三个位置创建汇总sheet
    ws_temp.cell(row=1, column=1, value='早自习数据汇总')
    for index in range(len(schoolProcessedDataFrame_list)):
        df_temp = schoolProcessedDataFrame_list[index]
        rows = dataframe_to_rows(df_temp)
        ws_temp.cell(row=2, column=1, value='学院')
        for r_idx, row in enumerate(rows, 1):
            if r_idx == 2:  # 第二行是一个空行，得忽略
                pass
            for c_idx, value in enumerate(row, 1):
                if c_idx == 2 and r_idx != 1:
                    ws_temp.cell(row=index + 3, column=1, value=value)  # 赋值学院
                if 15 >= c_idx >= 11:
                    if index == 1 and r_idx == 1:
                        ws_temp.cell(row=r_idx + 1, column=c_idx - 9, value=value)
                    else:
                        ws_temp.cell(row=index + 3, column=c_idx - 9, value=value)
                else:
                    pass
    wb.save(filePath)


if __name__ == '__main__':
    # 大概运行一次要4s多的样子
    startTime = datetime.datetime.now()
    filePath = r'C:\Users\HP\Desktop\standard_template.xlsx'
    ProcessOriginExcel(filePath)
    SetExcelStyle(filePath)
    endTime = datetime.datetime.now()
    print('程序运行时间为:', endTime-startTime)
